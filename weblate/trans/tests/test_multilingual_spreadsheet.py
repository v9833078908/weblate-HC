# Copyright © 2026 Weblate contributors
#
# SPDX-License-Identifier: GPL-3.0-or-later

from __future__ import annotations

import csv
from io import BytesIO, StringIO

from django.core.exceptions import ValidationError
from django.core.files.uploadedfile import SimpleUploadedFile
from openpyxl import load_workbook

from weblate.trans.tests.test_views import ViewTestCase


class MultilingualSpreadsheetExportTest(ViewTestCase):
    def create_component(self):
        return self.create_json()

    def test_csv_and_xlsx_export_one_component_table(self) -> None:
        from weblate.trans.multilingual_spreadsheet import (
            export_component,
            parse_upload,
        )

        unit = self.translation.unit_set.order_by("pk").first()
        assert unit is not None
        unit.translate(self.user, '=SUM(1,2), "Привет"\nnext', unit.state)

        expected_headers = (
            "key",
            *self.component.translation_set.order_by("language__code").values_list(
                "language__code", flat=True
            ),
        )
        csv_content = export_component(self.component, "csv")
        xlsx_content = export_component(self.component, "xlsx")

        self.assertEqual(
            parse_upload(
                self.component,
                SimpleUploadedFile(
                    "translations.csv", csv_content, content_type="text/csv"
                ),
            ).headers,
            expected_headers,
        )
        self.assertEqual(
            parse_upload(
                self.component,
                SimpleUploadedFile(
                    "translations.xlsx",
                    xlsx_content,
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                ),
            ).headers,
            expected_headers,
        )

        workbook = load_workbook(BytesIO(xlsx_content), data_only=False)
        self.assertEqual(len(workbook.worksheets), 1)
        self.assertEqual(
            workbook.active.cell(row=2, column=len(expected_headers)).data_type, "s"
        )

    def test_preview_rejects_reordered_placeholders(self) -> None:
        from weblate.trans.multilingual_spreadsheet import (
            build_preview,
            export_component,
            parse_upload,
        )

        unit = self.translation.unit_set.order_by("pk").first()
        assert unit is not None
        source_unit = unit.source_unit
        source_unit.source = "{0} {playerName}"
        source_unit.target = source_unit.source
        source_unit.save()
        unit.source = source_unit.source
        unit.save(update_fields=["source"])

        parsed = parse_upload(
            self.component,
            SimpleUploadedFile(
                "translations.csv", export_component(self.component, "csv")
            ),
        )
        values = list(parsed.rows[0].values)
        values[parsed.headers.index(self.translation.language.code)] = (
            "{playerName} {0}"
        )
        parsed = parsed.__class__(
            parsed.headers,
            (
                parsed.rows[0].__class__(parsed.rows[0].row_number, tuple(values)),
                *parsed.rows[1:],
            ),
        )

        with self.assertRaises(ValidationError):
            build_preview(self.component, parsed)

    def test_preview_accepts_translated_conditional_branches(self) -> None:
        from weblate.trans.multilingual_spreadsheet import (
            build_preview,
            export_component,
            parse_upload,
        )

        source = (
            "{hours:cond:>0?{hours}h. |}"
            "{minutes:cond:>0?{minutes}m. |}"
            "{seconds:cond:>=0?{seconds}s.|}"
        )
        target = (
            "{hours:cond:>0?{hours}Std. |}"
            "{minutes:cond:>0?{minutes}Min. |}"
            "{seconds:cond:>=0?{seconds}Sek.|}"
        )
        unit = self.translation.unit_set.order_by("pk").first()
        assert unit is not None
        source_unit = unit.source_unit
        source_unit.source = source
        source_unit.target = source
        source_unit.save()
        unit.source = source
        unit.save(update_fields=["source"])
        unit.translate(self.user, target, unit.state)

        parsed = parse_upload(
            self.component,
            SimpleUploadedFile(
                "translations.csv", export_component(self.component, "csv")
            ),
        )

        self.assertTrue(build_preview(self.component, parsed).changes)


class MultilingualSpreadsheetValidationTest(ViewTestCase):
    def create_component(self):
        return self.create_json()

    def _csv_rows(self) -> list[list[str]]:
        from weblate.trans.multilingual_spreadsheet import export_component

        return list(
            csv.reader(
                export_component(self.component, "csv").decode("utf-8").splitlines(),
                dialect="unix",
            )
        )

    def _parse_csv_rows(self, rows: list[list[str]]) -> None:
        from weblate.trans.multilingual_spreadsheet import parse_upload

        content = StringIO(newline="")
        csv.writer(content, dialect="unix").writerows(rows)
        parse_upload(
            self.component,
            SimpleUploadedFile("translations.csv", content.getvalue().encode("utf-8")),
        )

    def test_rejects_changed_component_schema_and_identity(self) -> None:
        rows = self._csv_rows()
        cases = (
            ("duplicate header", [rows[0][0], rows[0][0], *rows[0][2:]], rows[1:]),
            ("missing language", rows[0][:-1], [row[:-1] for row in rows[1:]]),
            ("unknown language", [*rows[0][:-1], "unknown"], rows[1:]),
            ("unknown key", rows[0], [["unknown", *rows[1][1:]], *rows[2:]]),
            ("duplicate key", rows[0], [rows[1], rows[1], *rows[2:]]),
        )
        for name, headers, body in cases:
            with self.subTest(name), self.assertRaises(ValidationError):
                self._parse_csv_rows([headers, *body])

    def test_rejects_xlsx_with_hidden_extra_worksheet(self) -> None:
        from weblate.trans.multilingual_spreadsheet import (
            export_component,
            parse_upload,
        )

        workbook = load_workbook(BytesIO(export_component(self.component, "xlsx")))
        worksheet = workbook.create_sheet("hidden")
        worksheet.sheet_state = "hidden"
        output = BytesIO()
        workbook.save(output)

        with self.assertRaises(ValidationError):
            parse_upload(
                self.component,
                SimpleUploadedFile("translations.xlsx", output.getvalue()),
            )

    def test_rejects_xlsx_with_inflated_dimension(self) -> None:
        # An attacker inflates the worksheet dimension (DoS / zip bomb) without
        # writing many real cells. In read-only mode openpyxl reports max_row
        # and max_column from the XML dimension attribute.
        import re
        import zipfile

        from weblate.trans.multilingual_spreadsheet import (
            export_component,
            parse_upload,
        )

        source = BytesIO(export_component(self.component, "xlsx"))
        output = BytesIO()
        with zipfile.ZipFile(source) as zin, zipfile.ZipFile(output, "w") as zout:
            for item in zin.infolist():
                payload = zin.read(item.filename)
                if item.filename == "xl/worksheets/sheet1.xml":
                    payload = re.sub(
                        rb'dimension ref="[^"]*"',
                        b'dimension ref="A1:XFD1048576"',
                        payload,
                        count=1,
                    )
                zout.writestr(item, payload)

        with self.assertRaises(ValidationError):
            parse_upload(
                self.component,
                SimpleUploadedFile("translations.xlsx", output.getvalue()),
            )

    def test_rejects_malformed_xlsx_xml(self) -> None:
        # Valid ZIP container, but the worksheet XML is truncated so openpyxl
        # raises InvalidFileException. Without the catch, that bubbles up as a 500.
        import zipfile

        from weblate.trans.multilingual_spreadsheet import (
            export_component,
            parse_upload,
        )

        source = BytesIO(export_component(self.component, "xlsx"))
        output = BytesIO()
        with zipfile.ZipFile(source) as zin, zipfile.ZipFile(output, "w") as zout:
            for item in zin.infolist():
                payload = zin.read(item.filename)
                if item.filename == "xl/worksheets/sheet1.xml":
                    payload = b"<worksheet><broken"
                zout.writestr(item, payload)

        with self.assertRaises(ValidationError):
            parse_upload(
                self.component,
                SimpleUploadedFile("translations.xlsx", output.getvalue()),
            )


class ComponentSpreadsheetImportDraftTest(ViewTestCase):
    def test_draft_is_owner_session_bound_and_expires(self) -> None:
        from django.utils import timezone

        from weblate.trans.models.multilingual_spreadsheet import (
            ComponentSpreadsheetImportDraft,
        )

        draft = ComponentSpreadsheetImportDraft.objects.create(
            owner=self.user,
            session_key=self.client.session.session_key,
            component=self.component,
            source_filename="translations.csv",
            preview_json="{}",
            baseline_json="{}",
        )

        self.assertIsNotNone(
            ComponentSpreadsheetImportDraft.get_active(
                token=draft.token,
                owner=self.user,
                session_key=self.client.session.session_key,
            )
        )
        draft.expires_at = timezone.now()
        draft.save(update_fields=["expires_at"])
        self.assertIsNone(
            ComponentSpreadsheetImportDraft.get_active(
                token=draft.token,
                owner=self.user,
                session_key=self.client.session.session_key,
            )
        )


class MultilingualSpreadsheetPluralTest(ViewTestCase):
    def test_plural_component_rejects_both_exports_and_upload(self) -> None:
        from weblate.trans.multilingual_spreadsheet import (
            export_component,
            parse_upload,
        )

        for format_name in ("csv", "xlsx"):
            with self.subTest(format_name), self.assertRaises(ValidationError):
                export_component(self.component, format_name)
        with self.assertRaises(ValidationError):
            parse_upload(
                self.component,
                SimpleUploadedFile(
                    "translations.csv", b"key,en,cs\nhello,hello,ahoj\n"
                ),
            )

    def test_plural_component_download_returns_error_redirect(self) -> None:
        from django.contrib.messages import get_messages
        from django.urls import reverse

        manager = self.user
        manager.is_superuser = True
        manager.save()
        self.client.force_login(manager)
        for format_name in ("csv", "xlsx"):
            with self.subTest(format_name):
                response = self.client.get(
                    reverse(
                        "multilingual-download",
                        kwargs={
                            "path": self.component.get_url_path(),
                            "format_name": format_name,
                        },
                    ),
                    follow=True,
                )
                self.assertEqual(response.status_code, 200)
                rendered = [
                    str(message) for message in get_messages(response.wsgi_request)
                ]
                self.assertTrue(
                    any("Plural units" in text for text in rendered),
                    rendered,
                )

    def test_preview_renders_table_and_cancel_button(self) -> None:
        import csv
        from io import StringIO

        from django.urls import reverse

        from weblate.trans.models.multilingual_spreadsheet import (
            ComponentSpreadsheetImportDraft,
        )
        from weblate.trans.multilingual_spreadsheet import export_component

        self.make_manager()
        self.user.clear_permissions_cache()
        self.client.force_login(self.user)
        self.component.source_translation.unit_set.all().delete()
        self.component.source_translation.add_unit(
            None, "ctx-a", "Hello", target="Hello", author=self.user
        )
        cs = self.component.translation_set.get(language__code="cs")
        cs_unit = cs.unit_set.get(context="ctx-a")
        cs_unit.translate(self.user, "Ahoj", 20)
        cs_unit.refresh_from_db()
        self.assertEqual(cs_unit.target, "Ahoj")

        export_bytes = export_component(self.component, "csv")
        rows = list(csv.reader(StringIO(export_bytes.decode("utf-8"))))
        headers = rows[0]
        source_code = self.component.source_language.code
        skip_indexes = {0}  # key column
        if "context" in headers:
            skip_indexes.add(headers.index("context"))
        if source_code in headers:
            skip_indexes.add(headers.index(source_code))
        edited_rows = [headers]
        for row in rows[1:]:
            new_row = list(row)
            for column_index, value in enumerate(new_row):
                if column_index in skip_indexes:
                    continue
                new_row[column_index] = value + " nov"
            edited_rows.append(new_row)
        buffer = StringIO()
        csv.writer(buffer).writerows(edited_rows)
        new_content = buffer.getvalue().encode("utf-8")

        uploaded = SimpleUploadedFile("translations.csv", new_content)
        response = self.client.post(
            reverse(
                "multilingual-upload",
                kwargs={"path": self.component.get_url_path()},
            ),
            data={"file": uploaded},
            follow=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '<th scope="col">en</th>')
        self.assertContains(response, '<th scope="col">cs</th>')
        self.assertContains(response, "Hello nov")
        self.assertContains(response, "/upload-multilingual/cancel/")
        self.assertContains(response, "/upload-multilingual/confirm/")

        draft = ComponentSpreadsheetImportDraft.objects.get(
            component=self.component, owner=self.user
        )
        cancel_response = self.client.post(
            reverse("multilingual-cancel", kwargs={"token": draft.token})
        )
        self.assertEqual(cancel_response.status_code, 302)
        self.assertFalse(
            ComponentSpreadsheetImportDraft.objects.filter(pk=draft.pk).exists()
        )


class ComponentSpreadsheetImportDraftCleanupTest(ViewTestCase):
    def test_cleanup_removes_expired_draft_and_private_file(self) -> None:
        from datetime import timedelta

        from django.utils import timezone

        from weblate.trans.models.multilingual_spreadsheet import (
            COMPONENT_SPREADSHEET_DRAFT_STORAGE,
            ComponentSpreadsheetImportDraft,
        )
        from weblate.trans.tasks import cleanup_component_spreadsheet_import_drafts

        draft = ComponentSpreadsheetImportDraft.objects.create(
            owner=self.user,
            session_key=self.client.session.session_key,
            component=self.component,
            source_filename="translations.csv",
            uploaded=SimpleUploadedFile("translations.csv", b"key,en,cs\n"),
            preview_json="{}",
            baseline_json="{}",
            expires_at=timezone.now() - timedelta(seconds=1),
        )
        name = draft.uploaded.name
        self.assertTrue(COMPONENT_SPREADSHEET_DRAFT_STORAGE.exists(name))

        cleanup_component_spreadsheet_import_drafts()

        self.assertFalse(
            ComponentSpreadsheetImportDraft.objects.filter(pk=draft.pk).exists()
        )
        self.assertFalse(COMPONENT_SPREADSHEET_DRAFT_STORAGE.exists(name))
