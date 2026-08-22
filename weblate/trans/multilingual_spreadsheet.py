# Copyright © 2026 Weblate contributors
#
# SPDX-License-Identifier: GPL-3.0-or-later

"""Component-level multilingual spreadsheet serialization and validation."""

from __future__ import annotations

import csv
from collections import Counter
from dataclasses import dataclass
from io import BytesIO, StringIO
from typing import TYPE_CHECKING, Literal
from zipfile import BadZipFile, ZipFile

from django.conf import settings
from django.core.exceptions import ValidationError

from weblate.formats.exporters import CSVExporter
from weblate.formats.external import CSV_DIALECT
from weblate.formats.ttkit import CSVUnit
from weblate.trans.protected_tokens import markup_tokens, placeholder_sequence
from weblate.trans.models import Unit
from weblate.utils.validators import validate_translation_upload_size
from weblate.utils.zip import ZipSafetyError, ZipSafetyLimits, validate_zip_members

if TYPE_CHECKING:
    from django.core.files.uploadedfile import UploadedFile

    from weblate.trans.models import Component


@dataclass(frozen=True, slots=True)
class SpreadsheetSchema:
    headers: tuple[str, ...]
    has_context: bool


@dataclass(frozen=True, slots=True)
class SpreadsheetRow:
    row_number: int
    values: tuple[str, ...]


@dataclass(frozen=True, slots=True)
class ParsedSpreadsheet:
    headers: tuple[str, ...]
    rows: tuple[SpreadsheetRow, ...]


@dataclass(frozen=True, slots=True)
class SpreadsheetPreview:
    parsed: ParsedSpreadsheet
    changes: tuple[SpreadsheetRow, ...]


def _error(message: str, *, row: int | None = None, column: str | None = None) -> None:
    if row is not None and column is not None:
        message = f"Row {row}, column {column}: {message}"
    raise ValidationError(message)


def _component_units(component: Component) -> tuple[list[Unit], dict[tuple[int, str], Unit]]:
    source_translation = component.source_translation
    source_units = list(source_translation.unit_set.order_by("pk"))
    source_ids = [unit.pk for unit in source_units]
    units = list(
        Unit.objects.filter(
            translation__component=component, source_unit_id__in=source_ids
        ).select_related("translation__language")
    )
    plural = next((unit for unit in units if unit.is_plural), None)
    if plural is not None:
        _error("Plural units cannot be exchanged in a multilingual spreadsheet.")
    return source_units, {
        (unit.source_unit_id, unit.translation.language.code): unit for unit in units
    }


def _schema(component: Component, source_units: list[Unit]) -> SpreadsheetSchema:
    languages = tuple(
        component.translation_set.order_by("language__code").values_list(
            "language__code", flat=True
        )
    )
    has_context = not component.has_template() and any(
        unit.context and unit.context != unit.source for unit in source_units
    )
    return SpreadsheetSchema(
        ("key", *languages, *(("context",) if has_context else ())), has_context
    )


def _identity(component: Component, unit: Unit, has_context: bool) -> tuple[str, ...]:
    key = unit.context if component.has_template() else unit.source
    return (key, unit.context) if has_context else (key,)


def _validate_identities(component: Component, source_units: list[Unit], has_context: bool) -> None:
    seen: set[tuple[str, ...]] = set()
    for row_number, unit in enumerate(source_units, start=2):
        identity = _identity(component, unit, has_context)
        if identity in seen:
            _error("Duplicate visible row identity.", row=row_number, column="key")
        seen.add(identity)


def _csv_value(value: str) -> str:
    return CSVExporter.string_filter(None, value)


def _serialize_csv(rows: list[tuple[str, ...]]) -> bytes:
    output = StringIO(newline="")
    writer = csv.writer(output, dialect=CSV_DIALECT)
    writer.writerows([tuple(_csv_value(value) for value in row) for row in rows])
    return output.getvalue().encode("utf-8")


def _serialize_xlsx(rows: list[tuple[str, ...]]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.cell.cell import TYPE_STRING

    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.title = "Weblate"
    for row_number, row in enumerate(rows, start=1):
        for column_number, value in enumerate(row, start=1):
            cell = worksheet.cell(row=row_number, column=column_number, value=value)
            cell.data_type = TYPE_STRING
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def export_component(component: Component, format_name: Literal["csv", "xlsx"]) -> bytes:
    source_units, units = _component_units(component)
    schema = _schema(component, source_units)
    _validate_identities(component, source_units, schema.has_context)
    rows = [schema.headers]
    languages = schema.headers[1 : -1 if schema.has_context else None]
    for source_unit in source_units:
        values = [*_identity(component, source_unit, schema.has_context)[:1]]
        values.extend(
            units[(source_unit.pk, language)].target
            for language in languages
        )
        if schema.has_context:
            values.append(source_unit.context)
        rows.append(tuple(values))
    if format_name == "csv":
        return _serialize_csv(rows)
    if format_name == "xlsx":
        return _serialize_xlsx(rows)
    msg = f"Unsupported multilingual spreadsheet format: {format_name}"
    raise ValueError(msg)


def _parse_csv(content: bytes) -> list[list[str]]:
    try:
        text = content.decode("utf-8")
    except UnicodeDecodeError as error:
        raise ValidationError("CSV must use UTF-8 encoding.") from error
    return [
        [CSVUnit.unescape_csv(value) for value in row]
        for row in csv.reader(StringIO(text), dialect=CSV_DIALECT)
    ]


def _parse_xlsx(content: bytes) -> list[list[str]]:
    try:
        with ZipFile(BytesIO(content)) as archive:
            validate_zip_members(
                archive,
                limits=ZipSafetyLimits(
                    max_members=1000,
                    max_total_uncompressed_size=settings.TRANSLATION_UPLOAD_MAX_SIZE,
                ),
            )
        from openpyxl import load_workbook

        workbook = load_workbook(BytesIO(content), data_only=False, read_only=True)
    except (BadZipFile, ZipSafetyError) as error:
        raise ValidationError("Invalid XLSX upload.") from error
    if len(workbook.worksheets) != 1:
        _error("XLSX upload must contain exactly one worksheet.")
    worksheet = workbook.worksheets[0]
    rows: list[list[str]] = []
    for row_number, row in enumerate(worksheet.iter_rows(), start=1):
        values: list[str] = []
        for cell in row:
            if cell.data_type == "f":
                _error("Formula cells are not supported.", row=row_number, column=cell.column_letter)
            values.append("" if cell.value is None else str(cell.value))
        if any(values):
            rows.append(values)
    return rows


def _validate_rows(component: Component, rows: list[list[str]]) -> ParsedSpreadsheet:
    source_units, _units = _component_units(component)
    schema = _schema(component, source_units)
    if not rows:
        _error("Spreadsheet is empty.")
    headers = tuple(rows[0])
    if len(headers) != len(set(headers)):
        _error("Spreadsheet contains duplicate headers.")
    if headers != schema.headers:
        _error("Spreadsheet headers do not match the component language schema.")
    _validate_identities(component, source_units, schema.has_context)
    source_by_identity = {
        _identity(component, source_unit, schema.has_context): source_unit
        for source_unit in source_units
    }
    parsed_rows: list[SpreadsheetRow] = []
    seen: set[tuple[str, ...]] = set()
    for row_number, row in enumerate(rows[1:], start=2):
        if len(row) != len(headers):
            _error("Unexpected number of cells.", row=row_number, column="key")
        identity = (row[0], row[-1]) if schema.has_context else (row[0],)
        if identity in seen:
            _error("Duplicate visible row identity.", row=row_number, column="key")
        seen.add(identity)
        if identity not in source_by_identity:
            _error("Unknown key or context.", row=row_number, column="key")
        parsed_rows.append(SpreadsheetRow(row_number, tuple(row)))
    if set(source_by_identity) != seen:
        _error("Spreadsheet is missing component rows.")
    return ParsedSpreadsheet(headers, tuple(parsed_rows))


def parse_upload(component: Component, uploaded: UploadedFile) -> ParsedSpreadsheet:
    validate_translation_upload_size(uploaded)
    uploaded.seek(0)
    content = uploaded.read()
    name = uploaded.name.lower()
    if name.endswith(".csv"):
        rows = _parse_csv(content)
    elif name.endswith(".xlsx"):
        rows = _parse_xlsx(content)
    else:
        _error("Only CSV and XLSX uploads are supported.")
    return _validate_rows(component, rows)


def build_preview(component: Component, parsed: ParsedSpreadsheet) -> SpreadsheetPreview:
    _validate_rows(component, [list(parsed.headers), *(list(row.values) for row in parsed.rows)])
    source_units, _units = _component_units(component)
    schema = _schema(component, source_units)
    source_by_identity = {
        _identity(component, source_unit, schema.has_context): source_unit
        for source_unit in source_units
    }
    source_code = component.source_language.code
    for row in parsed.rows:
        identity = (
            (row.values[0], row.values[-1]) if schema.has_context else (row.values[0],)
        )
        source = source_by_identity[identity].source
        for column, target in zip(parsed.headers[1:], row.values[1:], strict=True):
            if column == "context" or column == source_code or not target:
                continue
            if (
                Counter(markup_tokens(source)) != Counter(markup_tokens(target))
                or placeholder_sequence(source) != placeholder_sequence(target)
            ):
                _error(
                    "Protected tokens do not match the source.",
                    row=row.row_number,
                    column=column,
                )
    return SpreadsheetPreview(parsed, parsed.rows)
