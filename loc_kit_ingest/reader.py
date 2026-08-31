# Copyright © HCGameLoc
#
# SPDX-License-Identifier: GPL-3.0-or-later

from __future__ import annotations

import csv
from itertools import islice
from typing import TYPE_CHECKING

from loc_kit_ingest.langcode import language_code
from loc_kit_ingest.model import Diagnostic, Severity
from loc_kit_ingest.profile import RecordMapGrammar

if TYPE_CHECKING:
    from pathlib import Path

    from loc_kit_ingest.profile import ComponentProfile

# --------------------------------------------------------------------------- #
# Reading
# --------------------------------------------------------------------------- #

_CSV_DELIMITERS = (",", ";", "\t")
_DELIMITER_SCAN_ROWS = 20


class ReaderError(ValueError):
    """Fatal I/O or format error during reading."""


def read_sheets(path: Path) -> dict[str, list[list[str]]]:
    """
    Read a kit file into a dict of sheet_name -> list of rows.

    Each row is a list of strings. No trimming, no inference.
    CSV/TSV: one sheet named after the file stem.
    XLSX: one sheet per worksheet, using its name.
    """
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return {path.stem: _read_csv(path, _detect_delimiter(path))}
    if suffix == ".tsv":
        return {path.stem: _read_csv(path, "\t")}
    if suffix == ".xlsx":
        return _read_xlsx(path)
    msg = f"unsupported file suffix: {suffix!r}"
    raise ReaderError(msg)


def _read_csv(path: Path, delimiter: str) -> list[list[str]]:
    with path.open(newline="", encoding="utf-8-sig") as f:
        return [row for row in csv.reader(f, delimiter=delimiter)]


def _detect_delimiter(path: Path) -> str:
    """Choose the delimiter under which the sheet has a recognisable header."""
    best: tuple[int, int, int, str] | None = None
    for priority, delimiter in enumerate(_CSV_DELIMITERS):
        for index, row in enumerate(_scan_rows(path, delimiter)):
            hits = sum(1 for cell in row if language_code(cell) is not None)
            if hits:
                candidate = (index, -hits, priority, delimiter)
                if best is None or candidate < best:
                    best = candidate
                break
    return best[3] if best is not None else ","


def _scan_rows(path: Path, delimiter: str) -> list[list[str]]:
    """Parse the top of the file with one delimiter candidate."""
    try:
        with path.open(newline="", encoding="utf-8-sig") as f:
            return list(
                islice(csv.reader(f, delimiter=delimiter), _DELIMITER_SCAN_ROWS)
            )
    except csv.Error:
        return []


def _read_xlsx(path: Path) -> dict[str, list[list[str]]]:
    from openpyxl import load_workbook

    result: dict[str, list[list[str]]] = {}
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            rows: list[list[str]] = []
            for row in ws.iter_rows(values_only=True):
                rows.append([str(cell) if cell is not None else "" for cell in row])
            result[ws.title] = rows
    finally:
        wb.close()
    return result


# --------------------------------------------------------------------------- #
# Header validation
# --------------------------------------------------------------------------- #


def validate_sheet_headers(
    component: ComponentProfile, rows: list[list[str]]
) -> tuple[Diagnostic, ...]:
    """
    Check that every configured header cell matches the sheet's header row.

    Returns a tuple of ERROR diagnostics for each mismatch.
    """
    header_row = component.header_row  # 0-based
    if header_row >= len(rows):
        return (
            Diagnostic(
                Severity.ERROR,
                "header.missing",
                component.component,
                component.sheet,
                header_row + 1,
                f"header row {header_row + 1} does not exist in sheet",
            ),
        )

    header_cells = rows[header_row]

    diagnostics: list[Diagnostic] = []

    def _check(label: str, column: int, expected: str) -> None:
        if column >= len(header_cells):
            diagnostics.append(
                Diagnostic(
                    Severity.ERROR,
                    "header.mismatch",
                    component.component,
                    component.sheet,
                    header_row + 1,
                    f"column {column + 1} ({label}) is out of bounds "
                    f"(only {len(header_cells)} columns)",
                )
            )
            return
        actual = header_cells[column]
        if actual != expected:
            diagnostics.append(
                Diagnostic(
                    Severity.ERROR,
                    "header.mismatch",
                    component.component,
                    component.sheet,
                    header_row + 1,
                    f"header mismatch for {label}: "
                    f"expected {expected!r}, got {actual!r}",
                )
            )

    # Language columns
    for lang in component.languages:
        _check(f"language {lang.code}", lang.column, lang.header)

    # Key column (PO only)
    if component.key is not None:
        _check("key", component.key.column, component.key.header)

    # Comment columns
    for col in component.comments:
        _check(f"comment {col.name}", col.column, col.header)

    # Reference columns
    for col in component.references:
        _check(f"reference {col.name}", col.column, col.header)

    # Record-map section and note columns (profile v2)
    grammar = component.grammar
    if isinstance(grammar, RecordMapGrammar):
        if grammar.section_field is not None:
            _check(
                "section field",
                grammar.section_field.column,
                grammar.section_field.header,
            )
        for note in grammar.notes:
            label = (
                f"{note.scope} note"
                if note.language is None
                else f"{note.scope} note ({note.language})"
            )
            _check(label, note.column, note.header)
        if grammar.source_flags is not None:
            _check(
                "source flags",
                grammar.source_flags.column,
                grammar.source_flags.header,
            )
        for ignored in grammar.ignored_columns:
            _check("ignored column", ignored.column, ignored.header)

    return tuple(diagnostics)
