# Copyright © HCGameLoc
#
# SPDX-License-Identifier: GPL-3.0-or-later

r"""
Measure both glossary blind spots on a real localization kit.

Runs offline against an exported kit, without a live Weblate/Django instance,
so it stays usable on a kit before it is even imported. That is also why the
source-side exact matcher and the target-side exact test below are re-derived
from the same rules as the product rather than imported:

* the source-side matcher of ``weblate/glossary/models.py:220-222``, which needs
  the exact lowercased form with a non-word character or an edge on both sides;
* the target-side test of ``weblate/checks/glossary.py:53-59``, which needs
  ``term.target`` verbatim in the translation, with ``\\b`` boundaries unless the
  language is in ``NO_SPACE_LANGUAGES`` (weblate/lang/data.py:12).

The stem comparison, however, is imported directly from
``weblate.checks.morphology`` (Django-free): it is the one rule this probe
exists to validate, so it must never drift from what ``GlossaryCheck`` and the
source-side matcher actually run. See docs/misc/glossary_matcher_fingerprint.

Usage: python glossary-blindspot-probe.py <kit directory>
"""

from __future__ import annotations

import re
import sys
from collections import Counter
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from weblate.checks.morphology import (
    MORPHOLOGY_LANGUAGES,
    contains_inflected,
    get_algorithm,
    get_snowball_version,
)

# weblate/lang/data.py:12
NO_SPACE_LANGUAGES = {"zh", "ja", "th", "km", "lo", "my", "ko"}
NON_WORD_RE = re.compile(r"\W")
# Suffixes that make a boundary rejection an inflection rather than a chance
# substring hit. English source only; deliberately narrow.
INFLECTION_SUFFIXES = ("s", "es", "'s", "’s", "ed", "ing")


def load(path: Path) -> list[tuple[str, str]]:
    """Return (source, target) pairs of a Weblate xlsx export."""
    workbook = openpyxl.load_workbook(path, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    header = [str(cell or "") for cell in next(rows)]
    src, tgt = header.index("source"), header.index("target")
    pairs = [
        (str(row[src]), str(row[tgt]))
        for row in rows
        if row[src] is not None and row[tgt] is not None
    ]
    workbook.close()
    return pairs


def boundaries(text: str) -> set[int]:
    """Return word boundary offsets, as fetch_glossary_terms computes them."""
    result = {match.start() for match in NON_WORD_RE.finditer(text)}
    result.add(-1)
    result.add(len(text))
    return result


def occurrences(term: str, text: str) -> tuple[int, int]:
    """Return (accepted, boundary_rejected) occurrence counts of a term."""
    low, hits = text.lower(), boundaries(text.lower())
    accepted = rejected = 0
    start = low.find(term)
    while start != -1:
        end = start + len(term)
        if (start - 1 in hits) and (end in hits):
            accepted += 1
        elif low[end:].startswith(INFLECTION_SUFFIXES) and start - 1 in hits:
            rejected += 1
        start = low.find(term, start + 1)
    return accepted, rejected


def contains(term: str, text: str, *, boundary: bool) -> bool:
    """Test containment exactly as the check does."""
    escaped = re.escape(term)
    pattern = rf"\b{escaped}\b" if boundary else escaped
    return re.search(pattern, text, re.IGNORECASE) is not None


def main() -> int:
    kit = Path(sys.argv[1])
    glossaries = sorted(kit.glob("*glossary-*.xlsx"))
    languages = [path.stem.split("glossary-")[1] for path in glossaries]
    source_lang = "en"

    print(
        f"snowball {get_snowball_version()}, allowlist "
        f"{sorted(MORPHOLOGY_LANGUAGES)} (weblate.checks.morphology)"
    )

    terms_en = [
        source
        for source, _target in load(kit / f"space-arena-glossary-{source_lang}.xlsx")
    ]
    strings_en = [
        source
        for source, _target in load(
            kit / f"space-arena-game-strings-{source_lang}.xlsx"
        )
    ]
    print(f"glossary terms: {len(terms_en)}   source strings: {len(strings_en)}")

    # Source side: what the matcher sees, once, for the whole project.
    seen = Counter()
    missed = Counter()
    for term in terms_en:
        low = term.lower()
        for text in strings_en:
            got, lost = occurrences(low, text)
            seen[term] += got
            missed[term] += lost
    total_seen, total_missed = sum(seen.values()), sum(missed.values())
    print(
        f"source side: {total_seen} occurrences matched, "
        f"{total_missed} rejected as inflected "
        f"({total_missed * 100 // max(total_seen + total_missed, 1)}% invisible)"
    )
    worst = [(t, missed[t], seen[t]) for t in terms_en if missed[t]]
    worst.sort(key=lambda item: -item[1])
    for term, lost, got in worst[:8]:
        print(f"    {lost:5} invisible / {got:5} seen   {term!r}")

    # Target side: what the check reports today, and what it would report if the
    # containment test tolerated inflection (weblate.checks.morphology allowlist
    # only - a language absent from it, e.g. Indonesian, keeps its exact miss).
    print()
    print("  lang    strings  fires(exact)  fires(stemmed)  removed  copy-through")
    for lang in sorted(languages):
        if lang == source_lang:
            continue
        glossary = {
            source: target
            for source, target in load(kit / f"space-arena-glossary-{lang}.xlsx")
            if target.strip()
        }
        if not glossary:
            print(f"  {lang:7} glossary is untranslated, nothing to enforce")
            continue
        strings = load(kit / f"space-arena-game-strings-{lang}.xlsx")
        boundary = lang.split("_")[0] not in NO_SPACE_LANGUAGES
        copy_through = sum(1 for s, t in glossary.items() if s.strip() == t.strip())
        algorithm = get_algorithm(lang)

        fires = stemmed_fires = 0
        for source, target in strings:
            for term, expected in glossary.items():
                got, _lost = occurrences(term.lower(), source)
                if not got:
                    continue
                if contains(expected, target, boundary=boundary):
                    continue
                fires += 1
                if algorithm is not None and contains_inflected(expected, target, lang):
                    continue
                stemmed_fires += 1
        removed = fires - stemmed_fires
        print(
            f"  {lang:7} {len(strings):8} {fires:13} {stemmed_fires:15} "
            f"{removed:8} ({removed * 100 // max(fires, 1)}%) {copy_through:4}/{len(glossary)}"
        )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
